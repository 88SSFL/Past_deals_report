import pandas as pd
import numpy as np
from datetime import datetime
endereço_casa=r'C:\\Users\\lippe\\Downloads\\Banco de Dados Outstanding.xlsx'
endereço_trabalho=r'C:\\Users\\u3xp\\Downloads\\Banco de Dados Outstanding.xlsx'
saída_casa='C:\\Users\\lippe\\Documents\\output.xlsx'
saída_trabalho=r'C:\\Users\\u3xp\\Downloads\\Outstanding debts Petrobras.xlsx'

file_1 = pd.read_excel(endereço_trabalho,sheet_name='Tender')
file_2 = pd.read_excel(endereço_trabalho,sheet_name='Bancos')
file_3 = pd.read_excel(endereço_trabalho,sheet_name='OMR')

def dt_(string):
    if "/" in list(string)[:4]:
        d1,d2,m1,m2,y1,y2=0,2,3,5,6,10
    else:
        d1,d2,m1,m2,y1,y2=8,10,5,7,0,4
    return datetime.strptime(f'{string[d1:d2]}/{string[m1:m2]}/{string[y1:y2]}', '%d/%m/%Y')

maturity=list(map(lambda x:"".join(list(str(x))[-10:]) if "".join(list(str(x))[-4:])=="2115" else "".join(list(str(x))[-8:]),file_1 ['Ticker']))
maturity=list(map(lambda x:datetime.strptime(x, '%m/%d/%Y') if "".join(list(x)[-4:])=="2115" else datetime.strptime(x, '%m/%d/%y'),maturity))

for i in [6,7]:
    treatment=list()
    index=0
    for j in file_1.iloc[:,i]:
        if j>=0 and file_1.iloc[index,2] not in ["Captação", "Reabertura"]:
            treatment.append(-j)
        elif j<0 and file_1.iloc[index,2] in ["Captação", "Reabertura"]:
            treatment.append(-j)
        else:
            treatment.append(j)
        index+=1
    file_1.iloc[:,i]=treatment
for i in [3,4]:
    file_1.iloc[:,i]=list(map(lambda x: datetime.strptime(x, '%d/%m/%Y').date().strftime('%d/%m/%Y') if isinstance(x,str) else x.date().strftime('%d/%m/%Y'),file_1.iloc[:,i]))
col_1=file_1.columns
for i in col_1[-5:-2]:
    file_1[i]=file_1[i].fillna(0)
file_1['Fee (US$)']=file_1['Fee (US$)'].fillna(0)


file_3
#%%
#sheet_1_1
operation=file_1.iloc[:,0].unique()
operation=operation[:].astype(int)

launch, settlement, type_, ticker, principal, cash_flow, banks, fee, fee_pp=([] for i in range(9))

for i in operation:
    temporary=file_1.loc[file_1['Operação']==i].reset_index()
    type_.append(temporary.iloc[0,3])
    launch.append(temporary.iloc[0,4])
    settlement.append(temporary.iloc[0,5])
    ticker.append(temporary.iloc[:,2].nunique())
    principal.append(temporary.iloc[:,8].sum())
    cash_flow.append(-temporary.iloc[:,9].sum()+principal[-1])
    fee.append(temporary.iloc[0,11])
    fee_pp.append((abs(fee[-1]*100/principal[-1]) if fee[-1] not in ["Não aplicavel","Não Aplicavel"]  else "Não Aplicável"))
    banks.append(file_2[i].count())
sheet_1_1=pd.DataFrame(data={'Operação':operation, "Tipo": type_,'Data (anúncio)':launch, 'Data (liquidação)':settlement, "Nº de bancos": banks,'Ativos Envolvidos':ticker,'Principal (USD)': principal, 'Caixa (USD)': cash_flow,
                             'Fee':fee, "Fee %":list(map(lambda x: round(x,3) if isinstance (x,float) else x,fee_pp)), 'Fee_test':fee_pp})

sheet_1_1=sheet_1_1.sort_values(by='Operação').reset_index(drop=True)
fee_pp=sheet_1_1['Fee_test']
sheet_1_1=sheet_1_1.drop(columns=['Fee_test'])
sheet_1_1
#sheet_1_2
sheet_1_2=file_1.iloc[:,:8]
sheet_1_2['Caixa (USD)']=sheet_1_2.iloc[:,-1]-file_1.iloc[:,8]
maturity=list(map(lambda x:"".join(list(str(x))[-10:]) if "".join(list(str(x))[-4:])=="2115" else "".join(list(str(x))[-8:]),file_1 ['Ticker']))
maturity=list(map(lambda x:datetime.strptime(x, '%m/%d/%Y') if "".join(list(x)[-4:])=="2115" else datetime.strptime(x, '%m/%d/%y'),maturity))
col=sheet_1_2.columns
sheet_1_2['Maturity']=maturity
sheet_1_2=sheet_1_2.sort_values(by=['Operação','Maturity']).reset_index(drop=True).drop(columns=['Maturity']).rename(columns={'Principal\nMoeda Original':'Principal (Origem)',
                                                                                                                             'Principal\nUSD':'Principal (USD)','Ticker':"Ativo"})
sheet_1_2.head(60)
#%%
#sheet_1_3
operation_2,banks_2,fee_1_3,type_2=[],[],[],[]
for i in operation:
    temporary_2=file_2[file_2[i] == "Sim"].reset_index()
    operation_2.extend([i]*len(temporary_2.iloc[:,1]))
    if sheet_1_1.iloc[i-1,1] != "Make Whole":
        fee_1_3.extend([round(sheet_1_1.iloc[i-1,-2]/sheet_1_1.iloc[i-1,-6],2)]*len(temporary_2.iloc[:,1]))
    else:
        fee_1_3.extend(["Não Aplicável"]*len(temporary_2.iloc[:,1]))
    banks_2.extend(temporary_2.iloc[:,1])
    type_2.extend([sheet_1_1.iloc[i-1,1]]*len(temporary_2.iloc[:,1]))
sheet_1_3=pd.DataFrame(data={'Operação':operation_2,"Tipo":type_2,"Bancos": banks_2, 'Fee (USD)':list(map(lambda x: abs(x) if isinstance(x,float) else x ,fee_1_3))})

sheet_1_3=sheet_1_3.sort_values(by=['Operação','Bancos']).reset_index(drop=True)
sheet_1_3.head(60)
#%%
# sheet_2_1

asset = sheet_1_2.iloc[:, 1].unique()
deal_2_1, banks_3, currency, original_principal, usd_principal, usd_cash_flow, fee_2_1 = ([] for i in range(7))
original_principal_module, usd_principal_module, usd_cash_flow_module = [], [], []
for i in asset:
    temporary_2_1_1 = sheet_1_2.loc[sheet_1_2["Ativo"] == i].reset_index()
    deal_2_1.append(temporary_2_1_1.iloc[:, 1].nunique())
    temporary_2_1_2 = []
    for j in temporary_2_1_1.iloc[:, 1].unique():
        temporary_2_1_2.extend(sheet_1_3[sheet_1_3["Operação"] == j].iloc[:, 2])
    banks_3.append(len(set(temporary_2_1_2)))
    currency.append(temporary_2_1_1.iloc[0, 6])
    original_principal.append(temporary_2_1_1.iloc[:, 7].sum())
    usd_principal.append(temporary_2_1_1.iloc[:, 8].sum())
    usd_cash_flow.append(temporary_2_1_1.iloc[:, -1].sum())

    auxiliary_2_1_1, auxiliary_2_1_2, auxiliary_2_1_3 = 0, 0, 0

    for j, k, l in zip(temporary_2_1_1.iloc[:, 7], temporary_2_1_1.iloc[:, 8], temporary_2_1_1.iloc[:, 9]):
        auxiliary_2_1_1 += 0 if np.isnan(j) else abs(j)
        auxiliary_2_1_2 += 0 if np.isnan(k) else abs(k)
        auxiliary_2_1_3 += 0 if np.isnan(l) else abs(l)

    original_principal_module.append(auxiliary_2_1_1)
    usd_principal_module.append(auxiliary_2_1_2)
    usd_cash_flow_module.append(auxiliary_2_1_3)
sheet_2_1 = pd.DataFrame(data={'Ativo': asset, "Nº de Operações": deal_2_1, "Nº de bancos": banks_3, 'Moeda': currency,
                               'Principal (Origem)': original_principal,
                               'Principal (USD)': usd_principal, 'Caixa (USD)': usd_cash_flow,
                               "Volume Financeiro (Origem - Valor de Face)": original_principal_module,
                               "Volume Financeiro (USD - Valor de Face)": usd_principal_module,
                               "Volume Financeiro (USD - Fluxo de Caixa)": usd_cash_flow_module})
maturity = list(
    map(lambda x: "".join(list(str(x))[-10:]) if "".join(list(str(x))[-4:]) == "2115" else "".join(list(str(x))[-8:]),
        sheet_2_1['Ativo']))
maturity = list(
    map(lambda x: datetime.strptime(x, '%m/%d/%Y') if "".join(list(x)[-4:]) == "2115" else datetime.strptime(x,
                                                                                                             '%m/%d/%y'),
        maturity))
sheet_2_1['Maturity'] = maturity
sheet_2_1 = sheet_2_1.sort_values(by=['Maturity']).reset_index(drop=True).drop(columns=['Maturity'])

asset = sheet_2_1['Ativo']
#%%
# sheet_2_2

asset_2_2, type_2_2, launch_2_2, settlement_2_2, deal_2_2, original_principal_2_2, usd_principal_2_2, usd_cash_flow_2_2, fee_2_2 = (
[] for i in range(9))
for i in asset:
    temporary_2_2_1 = sheet_1_2.loc[sheet_1_2["Ativo"] == i].reset_index()
    deal_2_2.extend(temporary_2_2_1.iloc[:, 1])
    type_2_2.extend(temporary_2_2_1.iloc[:, 3])
    launch_2_2.extend(temporary_2_2_1.iloc[:, 4])
    settlement_2_2.extend(temporary_2_2_1.iloc[:, 5])
    asset_2_2.extend([i] * len(temporary_2_2_1.iloc[:, 1]))
    original_principal_2_2.extend(temporary_2_2_1.iloc[:, 7])
    usd_principal_2_2.extend(temporary_2_2_1.iloc[:, 8])
    usd_cash_flow_2_2.extend(temporary_2_2_1.iloc[:, -1])
    temporary_2_2_2 = []
    for j in range(len(temporary_2_2_1.iloc[:, 1])):
        temporary_2_2_2.append(
            fee_pp[temporary_2_2_1.iloc[j, 1] - 1] * abs(temporary_2_2_1.iloc[j, 7]) / 100 if isinstance(
                fee_pp[temporary_2_2_1.iloc[j, 1] - 1], float) else 0)

    fee_2_2.extend(temporary_2_2_2)
sheet_2_2 = pd.DataFrame(
    data={'Ativo': asset_2_2, " Operação": deal_2_2, 'Tipo de Operação': type_2_2, 'Data (anúncio)': launch_2_2,
          'Data (liquidação)': settlement_2_2, 'Principal (Origem)': original_principal_2_2,
          'Principal (USD)': usd_principal_2_2, 'Caixa (USD)': usd_cash_flow_2_2,
          'Fee (USD)': list(map(lambda x: round(x, 2), fee_2_2))})

sheet_2_2.head(60)

#sheet_2_1
fee_2_1,temporary_2_1_2=[],sheet_2_2.groupby('Ativo', as_index=False).agg({'Fee (USD)':'sum'})

for i in sheet_2_1['Ativo']:
    fee_2_1.append(temporary_2_1_2[temporary_2_1_2['Ativo']==i].iloc[0,1])
sheet_2_1['Fee (USD)']= fee_2_1

sheet_2_1.head(60)
#%%
# sheet_2_3,sheet_2_4,sheet_2_5,sheet_2_6


file_1['ano'] = list(map(lambda x: datetime.strptime(x, '%d/%m/%Y').year, file_1.iloc[:, 4]))
file_1['Fluxo de Caixa USD'] = file_1.iloc[:, 7] - file_1.iloc[:, 8]
sheet_2_suport = file_1.groupby(['Ticker', 'ano'], as_index=False).agg(
    {'Operação': 'count', 'Principal\nMoeda Original': 'sum',
     'Principal\nUSD': 'sum', 'Fluxo de Caixa USD': 'sum'})
year = sheet_2_suport.iloc[:, 1].unique()
year.sort()
asset_2_3, asset_2_4, asset_2_5, asset_2_6 = (asset for i in range(4))
asset_2 = [asset_2_3, asset_2_4, asset_2_5, asset_2_6]

for i in range(1, 4):
    asset_2[i] = np.append(asset_2[i], 'TOTAL')
sheet_2_3, sheet_2_4, sheet_2_5, sheet_2_6 = (pd.DataFrame(data={'Ativo': asset_2[i]}) for i in range(4))
sheet = [sheet_2_3, sheet_2_4, sheet_2_5, sheet_2_6]
for i in range(2, 6):
    for j in year:
        locals()[f'_{str(j)}_{i}'] = list()
        for k in asset:
            sheet_2_suport_1 = sheet_2_suport[(sheet_2_suport.iloc[:, 1] == j) & (sheet_2_suport.iloc[:, 0] == k)]
            locals()[f'_{str(j)}_{i}'].append(0 if sheet_2_suport_1.empty == True else sheet_2_suport_1.iloc[0, i])
        if i > 2:
            locals()[f'_{str(j)}_{i}'].append(sum(locals()[f'_{str(j)}_{i}']))
        sheet[i - 2][str(j)] = locals()[f'_{str(j)}_{i}']
        # sheet[i-2]['Total']=sheet[i-2].loc[:,list(map(str,year))].sum(axis = 1)

sheet_2_4
#%%
#sheet_2_7,sheet_2_8,sheet_2_9,sheet_2_10
file_3['Ano']=list(map(str,file_3['Data Liquidação']))
file_3['Ano']=list(map(lambda x:datetime.strptime(x[8:10]+"/"+x[5:7]+"/"+x[:4], '%d/%m/%Y').year,file_3['Ano']))
maturity_3=list(map(lambda x:"".join(list(str(x))[-10:]) if "".join(list(str(x))[-4:])=="2115" else "".join(list(str(x))[-8:]),file_3 ['Security Name']))
maturity_3=list(map(lambda x:datetime.strptime(x, '%m/%d/%Y') if "".join(list(x)[-4:])=="2115" else datetime.strptime(x, '%m/%d/%y'),maturity_3))
file_3['Maturity']=maturity_3
file_3=file_3.sort_values(by=['Maturity']).reset_index(drop=True)
file_3=file_3.drop(columns=['Maturity'])

sheet_2_suport_2=file_3.groupby(['Security Name','Ano'], as_index=False).agg({'Valor de Face':'sum','Valor Compra':'sum',
                                                              'Juros Acruados':'sum','Valor Total':'sum'})
asset_suport=file_3["Security Name"].unique()
year_2=sheet_2_suport_2.iloc[:,1].unique()
asset_2_7,asset_2_8,asset_2_9,asset_2_10=(asset_suport for i in range(4))
asset_3=[asset_2_7,asset_2_8,asset_2_9,asset_2_10]

for i in range(4):
    asset_3[i]=np.append(asset_3[i],'TOTAL')
sheet_2_7, sheet_2_8, sheet_2_9, sheet_2_10=(pd.DataFrame(data={'Ativo':asset_3[i]} )for i in range(4))
sheet_2=[sheet_2_7, sheet_2_8, sheet_2_9, sheet_2_10]
for i in range(7,11):
    for j in year_2:
        locals()[f'_{str(j)}_{i}']=list()
        for k in asset_suport:
            sheet_2_suport_3=sheet_2_suport_2[(sheet_2_suport_2.iloc[:,1]==j)&(sheet_2_suport_2.iloc[:,0]==k)]
            locals()[f'_{str(j)}_{i}'].append(0 if sheet_2_suport_3.empty==True else sheet_2_suport_3.iloc[0,i-5])
        locals()[f'_{str(j)}_{i}'].append(sum(locals()[f'_{str(j)}_{i}']))
        sheet_2[i-7][str(j)]=locals()[f'_{str(j)}_{i}']
    sheet_2[i-7]['Total']=sheet_2[i-7].loc[:,list(map(str,year_2))].sum(axis = 1)
sheet_2_7
#%%
### sheet_2_11,sheet_2_12
from dateutil.relativedelta import relativedelta
merger=[sheet_2_5,sheet_2_6,sheet_2_7,sheet_2_8]
for i in [11,12]:
    locals()[f'sheet_2_{i}']=pd.merge(merger[i-11],merger[i-9],how='outer').groupby('Ativo', as_index=False).sum()
    locals()[f'maturity_2_{i}']=list(map(lambda x:"".join(list(str(x))[-10:]) if "".join(list(str(x))[-4:])=="2115" else "".join(list(str(x))[-8:]),locals()[f'sheet_2_{i}'] ['Ativo'][:-1]))
    locals()[f'maturity_2_{i}']=list(map(lambda x:datetime.strptime(x, '%m/%d/%Y') if "".join(list(x)[-4:])=="2115" else datetime.strptime(x, '%m/%d/%y'),locals()[f'maturity_2_{i}']))
    locals()[f'maturity_2_{i}'].append(max(locals()[f'maturity_2_{i}'])+relativedelta(years=1))
    locals()[f'sheet_2_{i}']['Maturity']=locals()[f'maturity_2_{i}']
    locals()[f'sheet_2_{i}']=locals()[f'sheet_2_{i}'].sort_values(by=['Maturity']).reset_index(drop=True)
    locals()[f'sheet_2_{i}']=locals()[f'sheet_2_{i}'].drop(columns=['Maturity'])


#sheet_2_11.iloc[:,]
#%%
# sheet_3_1

for i in range(len(sheet_1_3.iloc[:, 3])):
    sheet_1_3.iloc[i, 3] = 0 if sheet_1_3.iloc[i, 3] == "Não Aplicável" else sheet_1_3.iloc[i, 3]
temporary_3_1_1 = sheet_1_3.groupby('Bancos', as_index=False).agg(
    {'Operação': 'count', 'Tipo': "count", 'Fee (USD)': 'sum'})
temporary_3_1_2 = sheet_1_2.groupby('Operação', as_index=False).sum()
banks_3_1 = temporary_3_1_1['Bancos']
n_deals_3_1 = temporary_3_1_1['Operação']
fee_3_1 = temporary_3_1_1['Fee (USD)']
original_principal_3_1, usd_principal_3_1, usd_cash_flow_3_1 = [], [], []
for i in range(len(banks_3_1)):
    temporary_3_1_3 = sheet_1_3[sheet_1_3['Bancos'] == banks_3_1[i]]
    deals_3_1 = list()
    deals_3_1.extend(temporary_3_1_3.iloc[:, 0])
    original_principal_3_1_1, usd_principal_3_1_1, usd_cash_flow_3_1_1 = [], [], []
    for j in deals_3_1:
        original_principal_3_1_1.append(temporary_3_1_2[temporary_3_1_2['Operação'] == j].iloc[0, 1])
        usd_principal_3_1_1.append(temporary_3_1_2[temporary_3_1_2['Operação'] == j].iloc[0, 2])
        usd_cash_flow_3_1_1.append(temporary_3_1_2[temporary_3_1_2['Operação'] == j].iloc[0, 3])

    original_principal_3_1_1 = list(map(abs, original_principal_3_1_1))
    usd_principal_3_1_1 = list(map(abs, usd_principal_3_1_1))
    usd_cash_flow_3_1_1 = list(map(abs, usd_cash_flow_3_1_1))
    original_principal_3_1.append(sum(original_principal_3_1_1))
    usd_principal_3_1.append(sum(usd_principal_3_1_1))
    usd_cash_flow_3_1.append(sum(usd_cash_flow_3_1_1))
sheet_3_1 = pd.DataFrame(
    data={'Bancos': banks_3_1, " Operação": n_deals_3_1, 'Principal (Origem)': original_principal_3_1,
          'Principal (USD)': usd_principal_3_1, 'Caixa (USD)': usd_cash_flow_3_1, 'Fee (USD)': fee_3_1})

sheet_3_1
#%%
#sheet_3_2

sheet_3_2=sheet_1_3.sort_values(by=['Bancos', 'Operação']).reset_index()
asset_3_2, launch_3_2, settlement_3_2, original_principal_3_2,usd_principal_3_2,usd_cash_flow_3_2=([] for i in range(6))
for i in sheet_3_2.iloc[:,1]:
    launch_3_2.append(sheet_1_1[sheet_1_1['Operação']==i].iloc[0,2])
    settlement_3_2.append(sheet_1_1[sheet_1_1['Operação']==i].iloc[0,3])
    asset_3_2.append(sheet_1_1[sheet_1_1['Operação']==i].iloc[0,5])
    original_principal_3_2.append(sheet_1_1[sheet_1_1['Operação']==i].iloc[0,6])
    usd_principal_3_2.append(sheet_1_1[sheet_1_1['Operação']==i].iloc[0,7])
    usd_cash_flow_3_2.append(sheet_1_1[sheet_1_1['Operação']==i].iloc[0,8])
sheet_3_2['Data de Anúncio']=launch_3_2
sheet_3_2['Data de Liquidação']=settlement_3_2
sheet_3_2['Ativos Envolvidos']=asset_3_2
sheet_3_2['Principal (Origem)']=original_principal_3_2
sheet_3_2['Principal (USD)']=usd_principal_3_2
sheet_3_2['Caixa (USD)']=usd_cash_flow_3_2
col=list(sheet_3_2.columns)
sheet_3_2=sheet_3_2[[col[3]]+col[1:3]+col[5:]+[col[4]]]
sheet_3_2
#%%
#sheet_3_3
year_3_3=list(set(list(map(lambda x:datetime.strptime(x, '%d/%m/%Y').year,settlement_3_2))))
year_3_3.sort()
sheet_3_2['Ano']=list(map(lambda x:datetime.strptime(x, '%d/%m/%Y').year,settlement_3_2))
sheet_3_suport=sheet_3_2.groupby(['Bancos', 'Ano'], as_index=False).agg({'Operação':'count','Ativos Envolvidos':'sum',
                                                                         'Principal (Origem)':'sum',
                                                              'Principal (USD)':'sum','Caixa (USD)':'sum','Fee (USD)':'sum'})
sheet_3_2=sheet_3_2.drop(columns=['Ano'])
sheet_3_suport.iloc[:,6]=list(map(lambda x: 0 if isinstance(x,str) else x, sheet_3_suport.iloc[:,6]))
banks_3_3,banks_3_4,banks_3_5,banks_3_6,banks_3_7, banks_3_8=(banks_3_1 for i in range(6))
banks_3=[banks_3_3,banks_3_4,banks_3_5,banks_3_6,banks_3_7, banks_3_8]
banks_3[5]=np.append(banks_3[5],'TOTAL')
sheet_3_3, sheet_3_4, sheet_3_5, sheet_3_6, sheet_3_7, sheet_3_8=(pd.DataFrame(data={'Bancos':banks_3[i]} )for i in range(6))
sheet_3=[sheet_3_3, sheet_3_4, sheet_3_5, sheet_3_6, sheet_3_7, sheet_3_8]
for i in range(2,8):
    for j in year_3_3:
        locals()[f'_{str(j)}_{i}']=list()
        for k in banks_3_1:
            sheet_3_suport_1=sheet_3_suport[(sheet_3_suport.iloc[:,1]==j)&(sheet_3_suport.iloc[:,0]==k)]
            locals()[f'_{str(j)}_{i}'].append(0 if sheet_3_suport_1.empty==True else sheet_3_suport_1.iloc[0,i])
        if i >6:

            locals()[f'_{str(j)}_{i}'].append(sum(locals()[f'_{str(j)}_{i}']))
        sheet_3[i-2][str(j)]=locals()[f'_{str(j)}_{i}']

sheet_3_4
#%%
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
import string

sheet = [sheet_1_1, sheet_1_2, sheet_1_3, sheet_2_1, sheet_2_2, sheet_2_3, sheet_2_4, sheet_2_5, sheet_2_6, sheet_2_7,
         sheet_2_8, sheet_2_9, sheet_2_10, sheet_2_11, sheet_2_12, sheet_3_1, sheet_3_2, sheet_3_3, sheet_3_4,
         sheet_3_5, sheet_3_6, sheet_3_7, sheet_3_8]
sheet_name = ['1.1', '1.2', '1.3', '2.1', '2.2', '2.3', '2.4', '2.5', '2.6', '2.7', '2.8', '2.9', '2.10', '2.11',
              '2.12', '3.1', '3.2', '3.3', '3.4', '3.5', '3.6', '3.7', '3.8']
sheet_description = ['Operações - Resumo', 'Operações - Detalhe', 'Operações - Bancos Participantes',
                     'Ativo - Resumo por Operação', 'Ativo - Detalhe por Operação', 'Ativo - Operações por Ano',
                     'Ativo - Valor de Face por Ano Ex OMR (Moeda Original)',
                     'Ativo - Valor de Face por Ano Ex OMR (US$)', 'Ativo - Valor de Face Módulo Ex OMR (US$)',
                     'Ativo - OMR - Valor de Face - Ano', 'Ativo - OMR - Valor de Compra - Ano',
                     'Ativo - OMR - Juros Acruados - Ano',
                     'Ativo - OMR - Valor Total - Ano', 'Ativo Ativo - Valor de Face por Ano Total',
                     'Ativo Ativo - Caixa por Ano Total',
                     'Bancos - Resumo por Operação', 'Bancos - Detalhes por Operação', 'Bancos - Operações por Ano',
                     'Bancos - Ativos Transacionados por Ano',
                     'Bancos - Valores de Face Transacionados por Ano (Moeda Original)',
                     'Bancos - Valores de Face Transacionados por Ano (US$)',
                     'Bancos - Caixa Transacionado por Ano (US$)', 'Bancos - Fees  por ano (US$)']



sheet_1_1.to_excel(saída_trabalho)
wb_0=load_workbook(saída_trabalho)
del wb_0[wb_0.sheetnames[0]] 
alphabet = list(string.ascii_uppercase)
col = []
with pd.ExcelWriter(saída_trabalho, engine='openpyxl') as writer:
    writer.book=wb_0
    for i, j in zip(sheet, sheet_name):
        i.to_excel(writer, sheet_name=j, startrow=3,index=False)
        ws=wb_0[wb_0.sheetnames[-1]]
        for k in range(len(i.columns)):
            if len(str(ws[get_column_letter(k+1)+str(5)].value))==10 and list(str(ws[get_column_letter(k+1)+str(5)].value))[2]=="/":
                for l in range(len(i)): 
                    ws[get_column_letter(k+1)+str(l+5)].number_format='dd/mm/yyyy'
                ws.column_dimensions[get_column_letter(k+1)].width=max(len(ws[get_column_letter(k+1)+str(4)].value)+3,12)

            elif str(ws[get_column_letter(k+1)+str(5)].value).isnumeric()==False and "." not in list(str(ws[get_column_letter(k+1)+str(5)].value)) :
                
                max_=len(ws[get_column_letter(k+1)+str(4)].value)
                for l in range(len(i)):
                    max_=max(max_,len(str(i.iloc[l,k])))
                ws.column_dimensions[get_column_letter(k+1)].width=max_+3
            elif "%" in list(ws[get_column_letter(k+1)+str(4)].value):
                max_=len(ws[get_column_letter(k+1)+str(4)].value)
                for l in range(len(i)):
                    try:
                        ws[get_column_letter(k+1)+str(l+5)].value=float(ws[get_column_letter(k+1)+str(l+5)].value)/100
                        max_=max(max_,len(list(str(ws[get_column_letter(k+1)+str(l+5)]))))
                    except:
                        ws[get_column_letter(k+1)+str(l+5)].value=ws[get_column_letter(k+1)+str(l+5)].value
                        max_=max(max_,len(list(str(ws[get_column_letter(k+1)+str(l+5)]))))
                    ws[get_column_letter(k+1)+str(l+5)].number_format='0.00%'
                ws.column_dimensions[get_column_letter(k+1)].width=max_+5        
            else:
                a=0
                while ws[get_column_letter(k+1)+str(5+a)].value in [0,"Não Aplicável"]:
                    a+=1                         
                if ws[get_column_letter(k+1)+str(5+a)].value==None or 1000>float(ws[get_column_letter(k+1)+str(5+a)].value)>-1000:
                    for l in range(len(i)): 
                        ws[get_column_letter(k+1)+str(l+5)].value=(int(ws[get_column_letter(k+1)+str(l+5)].value) if str(ws[get_column_letter(k+1)+str(l+5)].value).isnumeric() else ws[get_column_letter(k+1)+str(l+5)].value)
                        ws[get_column_letter(k+1)+str(l+5)].number_format='#,##0'
                    ws.column_dimensions[get_column_letter(k+1)].width=max(len(ws[get_column_letter(k+1)+str(4)].value)+2,10)             
                else: 
                    for l in range(len(i)): 
                        ws[get_column_letter(k+1)+str(l+5)].value=(float(ws[get_column_letter(k+1)+str(l+5)].value)if str(ws[get_column_letter(k+1)+str(l+5)].value).isnumeric() else ws[get_column_letter(k+1)+str(l+5)].value)
                        ws[get_column_letter(k+1)+str(l+5)].number_format='#,##0.00'
                    max_=len(ws[get_column_letter(k+1)+str(4)].value)+2
                    for m in range(len(i)):
                        try:
                            max_=max(max_,len(str(i.iloc[m,k]*100//1)))
                        except:
                            max_=max(max_,len(i.iloc[m,k]))
                    ws.column_dimensions[get_column_letter(k+1)].width=max_+5                   

    for i in sheet:
        col.extend(i.columns)
    col = set(col)

    indice = wb_0.create_sheet("Índice", 0)
    indice['b2'] = "Índice"
    indice['b2'].font = Font(size=14, bold=True)
    indice.sheet_view.showGridLines = False
    ft = Font(size=20, bold=True)
    
    for i in range(len(sheet)):

        indice[f'b{3 + i}'] = f'=HYPERLINK("[Outstanding debts Petrobras.xlsx]{sheet_name[i]}!A1","{sheet_name[i]}")'
        indice[f'b{3 + i}'].style = "Hyperlink"
        indice[f'c{3 + i}'] = sheet_description[i]

        aba = wb_0[sheet_name[i]]
        for j in range(sheet[i].shape[1] + 1):
            if sheet_name[i][0] == '1':
                aba.sheet_properties.tabColor = "FF6347"
            elif sheet_name[i][0] == '2':
                aba.sheet_properties.tabColor = "008000"
            else:
                aba.sheet_properties.tabColor = "4682B4"
        aba['A1']='=HYPERLINK("[Outstanding debts Petrobras.xlsx]Índice!A1","Retorno")'
        aba['A1'].style = "Hyperlink"
        aba['B2'] = sheet_description[i]
        aba['B2'].font = ft
        aba.sheet_view.showGridLines = False

    wb_0.save('saída_trabalho')
# sheet.cell(row = 1, column = 1).value = ' hello '

# sheet.cell(row = 2, column = 2).value = ' everyone '
# sheet.row_dimensions[1].height = 70
# sheet.column_dimensions['B'].width = 20
# wb.save('dimension.xlsx')